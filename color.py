#表の先頭のセルに色を付けるファイル
import openpyxl
from openpyxl.styles import PatternFill

#excelファイルの読み込み
wb = openpyxl.load_workbook('/Users/maceponta/Downloads/ponta.xlsx')
# 2回繰り返す
for counter in range(2):
 #1回目：シート1、2回目:シート2
 ws = wb[f'test{counter + 1}']
 #行データを取得して１行目のセルを黄色に変換
 for row in ws.iter_rows():
        for cell in row:
            if cell.row == 1:
                cell.fill = PatternFill(fgColor='FFFF00',bgColor="FFFF00", fill_type = "solid")
#excelファイルを保存する
wb.save('/Users/maceponta/Downloads/ponta.xlsx')